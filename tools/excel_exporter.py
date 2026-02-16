"""
Excel Exporter for the AI Intelligence System.
Exports all findings to a structured Excel workbook with multiple sheets.
"""

import os
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import Database

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".tmp")

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
GAP_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
WEEKEND_FILL = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
HIGH_SCORE_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, num_cols):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws, min_width=10, max_width=60):
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            # Use first line only for width calculation
            first_line = val.split("\n")[0] if val else ""
            lengths.append(len(first_line))
        best = max(lengths) if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best + 2, max_width))


def _parse_json_field(value):
    """Safely parse a JSON string field."""
    if not value:
        return []
    if isinstance(value, list):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else [str(parsed)]
    except (json.JSONDecodeError, TypeError):
        return [str(value)]


def build_ideas_sheet(wb, db):
    """Sheet 1: All scored ideas — the main findings view."""
    ws = wb.active
    ws.title = "Ideas"

    headers = [
        "Rank", "Title", "Description", "Momentum", "Buildability",
        "Urgency", "Combined", "Effort", "Gap Pattern",
        "Source Subreddits", "Created"
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    ideas = db.get_ideas(min_score=0.0, limit=500)
    for rank, idea in enumerate(ideas, 1):
        subs = ", ".join(_parse_json_field(idea["source_subreddits"]))
        row = [
            rank,
            idea["title"],
            idea["description"][:300] if idea["description"] else "",
            idea["momentum_score"],
            idea["buildability_score"],
            idea["urgency_score"],
            round(idea["combined_score"], 2),
            idea["effort_estimate"],
            "YES" if idea["gap_pattern"] else "",
            subs,
            idea["created_at"],
        ]
        ws.append(row)

        row_num = rank + 1
        # Highlight gap patterns
        if idea["gap_pattern"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = GAP_FILL
        # Highlight weekend projects
        elif idea["effort_estimate"] == "weekend":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = WEEKEND_FILL
        # Highlight high-scoring ideas
        elif idea["combined_score"] >= 5.0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = HIGH_SCORE_FILL

        # Wrap text for description
        ws.cell(row=row_num, column=3).alignment = WRAP_ALIGNMENT

    _auto_width(ws)
    # Fix description column wider
    ws.column_dimensions["C"].width = 50


def build_weekend_projects_sheet(wb, db):
    """Sheet 2: Weekend projects — quick wins."""
    ws = wb.create_sheet("Weekend Projects")

    headers = [
        "Rank", "Title", "Buildability", "Momentum", "Urgency",
        "Combined", "Source Subreddits", "Description"
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    ideas = db.get_ideas(min_score=0.0, limit=500)
    weekend = [i for i in ideas if i["effort_estimate"] == "weekend"]
    weekend.sort(key=lambda x: x["combined_score"], reverse=True)

    for rank, idea in enumerate(weekend, 1):
        subs = ", ".join(_parse_json_field(idea["source_subreddits"]))
        ws.append([
            rank,
            idea["title"],
            idea["buildability_score"],
            idea["momentum_score"],
            idea["urgency_score"],
            round(idea["combined_score"], 2),
            subs,
            idea["description"][:300] if idea["description"] else "",
        ])
        ws.cell(row=rank + 1, column=8).alignment = WRAP_ALIGNMENT

    _auto_width(ws)
    ws.column_dimensions["H"].width = 50


def build_gap_patterns_sheet(wb, db):
    """Sheet 3: Gap patterns — unsolved problems."""
    ws = wb.create_sheet("Gap Patterns")

    headers = [
        "Rank", "Title", "Urgency", "Momentum", "Combined",
        "Source Subreddits", "Description"
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    gaps = db.get_gap_patterns(limit=100)
    for rank, idea in enumerate(gaps, 1):
        subs = ", ".join(_parse_json_field(idea["source_subreddits"]))
        ws.append([
            rank,
            idea["title"],
            idea["urgency_score"],
            idea["momentum_score"],
            round(idea["combined_score"], 2),
            subs,
            idea["description"][:300] if idea["description"] else "",
        ])
        ws.cell(row=rank + 1, column=7).alignment = WRAP_ALIGNMENT

    if not gaps:
        ws.append(["", "No gap patterns detected this cycle"])

    _auto_width(ws)
    ws.column_dimensions["G"].width = 50


def build_insights_sheet(wb, db):
    """Sheet 4: Raw insights — full detail."""
    ws = wb.create_sheet("Insights")

    headers = [
        "ID", "Type", "Summary", "Entities", "Sentiment",
        "Urgency Signals", "Confidence", "Source Post",
        "Subreddit", "Post Score", "Comments", "Created"
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    insights = db.get_insights(days=30)
    for ins in insights:
        entities = ", ".join(_parse_json_field(ins["entities"]))
        signals = ", ".join(_parse_json_field(ins["urgency_signals"]))
        ws.append([
            ins["id"],
            ins["insight_type"],
            ins["summary"],
            entities,
            round(ins["sentiment"], 2) if ins["sentiment"] else 0,
            signals,
            round(ins["confidence"], 2) if ins["confidence"] else 0,
            ins["content_title"],
            ins["source_name"],
            ins["post_score"],
            ins["num_comments"],
            ins["created_at"],
        ])

    _auto_width(ws)
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30


def build_content_sheet(wb, db):
    """Sheet 5: Collected posts — source data."""
    ws = wb.create_sheet("Collected Posts")

    headers = [
        "ID", "Subreddit", "Title", "Author", "Score",
        "Comments", "Engagement", "URL", "Posted At", "Processed"
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    content = db.conn.execute(
        """SELECT c.*, s.name as source_name
           FROM content c JOIN sources s ON c.source_id = s.id
           ORDER BY c.engagement_score DESC"""
    ).fetchall()

    for post in content:
        ws.append([
            post["id"],
            post["source_name"],
            post["title"],
            post["author"],
            post["score"],
            post["num_comments"],
            round(post["engagement_score"], 2),
            post["url"],
            post["posted_at"],
            "Yes" if post["processed"] else "No",
        ])

    _auto_width(ws)
    ws.column_dimensions["C"].width = 50


def build_summary_sheet(wb, db):
    """Sheet 6: Dashboard summary stats."""
    ws = wb.create_sheet("Summary")

    # Move to first position
    wb.move_sheet(ws, offset=-5)

    stats = db.get_stats()
    sources = db.get_sources()

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"AI Intelligence Report - {datetime.now().strftime('%Y-%m-%d')}"
    title_cell.font = Font(bold=True, size=16, color="1a1a2e")

    # Stats section
    ws["A3"] = "Pipeline Stats"
    ws["A3"].font = Font(bold=True, size=13)

    stat_rows = [
        ("Total Posts Collected", stats["content"]),
        ("Insights Extracted", stats["insights"]),
        ("Ideas Scored", stats["ideas"]),
        ("Unprocessed Posts", stats["unprocessed_content"]),
    ]
    for i, (label, value) in enumerate(stat_rows, 4):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value

    # Effort breakdown
    ideas = db.get_ideas(min_score=0.0, limit=500)
    effort_counts = {}
    for idea in ideas:
        e = idea["effort_estimate"]
        effort_counts[e] = effort_counts.get(e, 0) + 1

    row = 9
    ws[f"A{row}"] = "Ideas by Effort"
    ws[f"A{row}"].font = Font(bold=True, size=13)
    row += 1
    for effort, count in sorted(effort_counts.items(), key=lambda x: -x[1]):
        ws[f"A{row}"] = effort.capitalize()
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = count
        row += 1

    # Gap patterns count
    gap_count = sum(1 for i in ideas if i["gap_pattern"])
    row += 1
    ws[f"A{row}"] = "Gap Patterns Detected"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = gap_count

    # Subreddit stats
    row += 2
    ws[f"A{row}"] = "Tracked Subreddits"
    ws[f"A{row}"].font = Font(bold=True, size=13)
    row += 1
    ws[f"A{row}"] = "Subreddit"
    ws[f"B{row}"] = "Authority"
    ws[f"C{row}"] = "Last Collected"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)
    ws[f"C{row}"].font = Font(bold=True)
    row += 1
    for src in sources:
        ws[f"A{row}"] = src["name"]
        ws[f"B{row}"] = src["authority_score"]
        ws[f"C{row}"] = src["last_collected_at"] or "never"
        row += 1

    _auto_width(ws)


def export_to_excel():
    """Export all findings to a structured Excel workbook."""
    db = Database()
    wb = Workbook()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Building Excel workbook...")

    build_ideas_sheet(wb, db)
    print("  - Ideas sheet")

    build_weekend_projects_sheet(wb, db)
    print("  - Weekend Projects sheet")

    build_gap_patterns_sheet(wb, db)
    print("  - Gap Patterns sheet")

    build_insights_sheet(wb, db)
    print("  - Insights sheet")

    build_content_sheet(wb, db)
    print("  - Collected Posts sheet")

    build_summary_sheet(wb, db)
    print("  - Summary sheet")

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"intelligence_report_{date_str}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # If file is locked (e.g. open in Excel or OneDrive sync), use timestamped name
    try:
        wb.save(filepath)
    except PermissionError:
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"intelligence_report_{ts}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
    db.close()

    print(f"\nExported to: {filepath}")
    return filepath


if __name__ == "__main__":
    export_to_excel()
